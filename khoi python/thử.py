from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = Workbook()
ws = wb.active

# Gộp ô tiêu đề
ws.merge_cells("A1:D1")
ws["A1"] = "BÁO CÁO HỌC SINH"

# Font tiêu đề
ws["A1"].font = Font(name="Times New Roman", size=16, bold=True, color="FFFFFF")

# Căn giữa
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

# Tô nền
ws["A1"].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

# Viền
thin = Side(border_style="thin", color="000000")
ws["A1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)

# Chiều cao hàng 1
ws.row_dimensions[1].height = 30

wb.save("tieu_de.xlsx")
