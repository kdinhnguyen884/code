from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

wb = Workbook()
ws = wb.active

# ==========================================================
# ====================== TIÊU ĐỀ CHÍNH ======================
# ==========================================================

ws.merge_cells("A1:D1")
ws["A1"] = "BÁO CÁO HỌC SINH"
ws.row_dimensions[1].height = 25

# ✔ Tạo font mới (KHÔNG dùng copy)
ws["A1"].font = Font(name="Times New Roman", size=14, bold=True, color="FFFFFF")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

ws["A1"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# ==========================================================
# ===================== TIÊU ĐỀ CỘT =========================
# ==========================================================

ws["A2"] = "Họ và tên"
ws["B2"] = "Tuổi"
ws["C2"] = "Lớp học"
ws["D2"] = "sở thích"

header_font = Font(name="Times New Roman", size=12, bold=True, color="0000FF")

for cell in ws[2]:
    cell.font = header_font

# ==========================================================
# ======================== DỮ LIỆU ==========================
# ==========================================================

ws.append(["Khôi", 25, "6D","bơi"])
ws.append(["Lan", 22, "7A","cầu lông"])

# ==========================================================
# ================ CĂN GIỮA TOÀN BỘ BẢNG ====================
# ==========================================================

align_center = Alignment(horizontal="center", vertical="center")

for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                        min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = align_center

# ==========================================================
# ======================== VIỀN Ô ============================
# ==========================================================

thin = Side(border_style="thin")
border = Border(top=thin, bottom=thin, left=thin, right=thin)

for row in ws["A1:D4"]:
    for cell in row:
        cell.border = border

# ==========================================================
# ======= TỰ ĐỘNG ĐIỀU CHỈNH ĐỘ RỘNG CỘT (AUTO FIT) =========
# ==========================================================

for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)

    for cell in col:
        try:
            value_length = len(str(cell.value))
            if value_length > max_length:
                max_length = value_length
        except:
            pass

    ws.column_dimensions[col_letter].width = max_length + 2

ws.title = "Báo cáo"

# ==========================================================
# =========== COPY SANG SHEET MỚI KHÔNG MẤT STYLE ===========
# ==========================================================

ws1 = ws
ws2 = wb.create_sheet("Danh sách lớp")

# Copy nội dung + style
for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row,
                         min_col=1, max_col=ws1.max_column):
    for cell in row:
        new_cell = ws2[cell.coordinate]
        new_cell.value = cell.value

        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = cell.number_format
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)

# Copy merged cells
for merged in ws1.merged_cells.ranges:
    ws2.merge_cells(str(merged))

# Copy column width
for col, dim in ws1.column_dimensions.items():
    ws2.column_dimensions[col].width = dim.width

# Copy row height
for row, dim in ws1.row_dimensions.items():
    ws2.row_dimensions[row].height = dim.height

# ==========================================================
# ======================== LƯU FILE ==========================
# ==========================================================

wb.save("du_lieu_moi.xlsx")
